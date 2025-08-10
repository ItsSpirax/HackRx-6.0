import base64
import concurrent.futures
import csv
import email
import hashlib
import io
import os
import re
import subprocess
import tempfile
from pathlib import Path
from typing import List
from urllib.parse import urlparse, unquote

import docx
import fitz
import redis
import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from langchain.text_splitter import RecursiveCharacterTextSplitter
from mistralai import Mistral
from openpyxl import load_workbook
from pdf2image import convert_from_path
from PIL import Image
from transformers import AutoTokenizer

load_dotenv()

# Create directory for storing document data
document_data_dir = Path("document_data")
document_data_dir.mkdir(exist_ok=True)

# Initialize Redis client for caching document metadata
client = redis.Redis(
    host=os.getenv("REDIS_HOST"),
    port=int(os.getenv("REDIS_PORT")),
    db=int(os.getenv("REDIS_DB")),
    decode_responses=True,
)

# Load tokenizer for precise token counting during chunking
tokenizer = AutoTokenizer.from_pretrained(
    os.getenv("TOKENIZER_MODEL"), use_fast=True, token=os.getenv("HF_API_KEY")
)


def identify_document_type(url: str) -> str:
    """
    Detect document type from URL extension and HTTP headers
    Returns document type string for appropriate processing pipeline
    """
    extension_map = {
        "pdf": "pdf",
        "docx": "docx",
        "eml": "email",
        "msg": "email",
        "jpg": "jpg",
        "jpeg": "jpg",
        "png": "png",
        "xlsx": "excel",
        "xls": "excel",
        "pptx": "ppt",
        "ppt": "ppt",
        "csv": "csv",
        "htm": "html",
        "html": "html",
    }

    def get_extension_from_url(url: str) -> str:
        """Extract file extension from URL path"""
        parsed_path = urlparse(url).path.lower()
        ext = Path(parsed_path).suffix.lower().lstrip(".")
        return extension_map.get(ext)

    try:
        # Check HTTP headers for content type information
        head = requests.head(url, allow_redirects=True, timeout=5)
        content_type = head.headers.get("Content-Type", "").lower()
        content_disp = head.headers.get("Content-Disposition", "").lower()
        parsed_path = urlparse(url).path.lower()

        # Extract filename from headers or URL
        filename = None
        if "filename=" in content_disp:
            filename = unquote(content_disp.split("filename=")[-1].strip('"; '))
        if not filename:
            filename = os.path.basename(parsed_path)

        # Check file extension first
        ext = Path(filename).suffix.lower().lstrip(".")
        if ext in extension_map:
            return extension_map[ext]

        # Fallback to content type detection
        if "pdf" in content_type:
            return "pdf"
        elif "word" in content_type:
            return "docx"
        elif "excel" in content_type or "spreadsheet" in content_type:
            return "excel"
        elif "powerpoint" in content_type:
            return "ppt"
        elif "image/jpeg" in content_type:
            return "jpg"
        elif "image/png" in content_type:
            return "png"
        elif "message/rfc822" in content_type or "eml" in content_type:
            return "email"
        elif "text/csv" in content_type or "application/csv" in content_type:
            return "csv"
        elif "text/html" in content_type or "application/xhtml+xml" in content_type:
            return "html"

        return None

    except Exception:
        return get_extension_from_url(url)


def download_document(url: str) -> bytes:
    """
    Download document content from URL with proper error handling
    Returns raw bytes for further processing
    """
    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        return response.content
    except requests.RequestException as e:
        raise Exception(f"Failed to download document: {e}")


def parse_document_content(content: bytes, doc_type: str, url: str) -> str:
    """
    Extract text content from various document formats
    Uses format-specific parsers and OCR for images/presentations
    """

    def clean_text(text: str) -> str:
        """Clean and normalize extracted text"""
        text = re.sub(r"<.*?>", " ", text)  # Remove HTML tags
        text = re.sub(r"[\n\r]+", " ", text)  # Normalize line breaks
        text = re.sub(r"\s+", " ", text)  # Normalize whitespace
        text = re.sub(r"(\. ?){2,}", ". ", text)  # Fix multiple periods
        text = re.sub(r"!\[.*?\]\(.*?\)", "", text)  # Remove markdown images
        return text.strip()

    def extract_pdf_text(content: bytes) -> str:
        with fitz.open(stream=content, filetype="pdf") as doc:
            texts = [clean_text(page.get_text() or "") for page in doc]
        return " ".join(texts)

    def extract_docx_text(content: bytes) -> str:
        doc = docx.Document(io.BytesIO(content))
        sections = []
        current = []

        for para in doc.paragraphs:
            text = clean_text(para.text)
            if not text:
                continue

            style = para.style.name.lower() if para.style else ""
            if "heading" in style:
                if current:
                    sections.append(" ".join(current))
                    current = []
                current.append(f"## {text}")
            elif "list" in style:
                current.append(f"- {text}")
            else:
                current.append(text)

        if current:
            sections.append(" ".join(current))

        return clean_text(" ".join(sections))

    def extract_email_text(content: bytes) -> str:
        msg = email.message_from_bytes(content)
        parts = []

        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/plain":
                    try:
                        text = part.get_payload(decode=True).decode("utf-8", "ignore")
                        parts.append(clean_text(text))
                    except Exception:
                        continue
        else:
            try:
                text = msg.get_payload(decode=True).decode("utf-8", "ignore")
                parts.append(clean_text(text))
            except Exception:
                pass

        return clean_text(" ".join(parts))

    def extract_excel_text(content: bytes) -> str:
        wb = load_workbook(filename=io.BytesIO(content), data_only=True)
        texts = []

        for sheet in wb.worksheets:
            sheet_name = sheet.title
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue

            header_row = next((r for r in rows if any(r)), None)
            if not header_row:
                continue
            headers = [
                str(cell).strip() if cell else f"Column{i+1}"
                for i, cell in enumerate(header_row)
            ]

            for row_idx, row in enumerate(rows[1:], start=2):
                if not any(row):
                    continue

                cells = []
                for i, cell in enumerate(row):
                    key = headers[i] if i < len(headers) else f"Column{i+1}"
                    value = str(cell).strip()
                    if value:
                        cells.append(f"{key}: {value}")

                if cells:
                    context_row = (
                        f"[Sheet: {sheet_name}, Row: {row_idx}] " + " | ".join(cells)
                    )
                    texts.append(context_row)

        return clean_text(" ".join(texts))

    def extract_with_mistral_ocr(content: bytes, doc_type: str, url: str = None) -> str:
        client = Mistral(api_key=os.getenv("MISTRAL_API_KEY"))

        try:
            if doc_type in {"jpg", "jpeg", "png"}:
                if not url:
                    raise ValueError("Image URL required for image OCR")
                ocr_response = client.ocr.process(
                    model="mistral-ocr-latest",
                    document={"type": "image_url", "image_url": url},
                    include_image_base64=False,
                )
                pages = ocr_response.pages

            elif doc_type in {"ppt", "pptx"}:
                with tempfile.TemporaryDirectory() as tmpdir:
                    pptx_path = Path(tmpdir) / "slides.pptx"
                    pdf_path = Path(tmpdir) / "slides.pdf"

                    with open(pptx_path, "wb") as f:
                        f.write(content)

                    libreoffice_profile = Path("/tmp/libreprofile")
                    libreoffice_profile.mkdir(parents=True, exist_ok=True)

                    env = os.environ.copy()
                    env["HOME"] = "/home/hackrx"

                    subprocess.run(
                        [
                            "libreoffice",
                            "--headless",
                            f"-env:UserInstallation=file://{libreoffice_profile}",
                            "--convert-to",
                            "pdf",
                            "--outdir",
                            str(tmpdir),
                            str(pptx_path),
                        ],
                        check=True,
                        env=env,
                        stdout=subprocess.PIPE,
                        stderr=subprocess.PIPE,
                    )

                    images = convert_from_path(str(pdf_path))

                    def ocr_image(img: Image.Image) -> str:
                        buffered = io.BytesIO()
                        img.save(buffered, format="JPEG")
                        base64_img = base64.b64encode(buffered.getvalue()).decode(
                            "utf-8"
                        )

                        ocr_response = client.ocr.process(
                            model="mistral-ocr-latest",
                            document={
                                "type": "image_url",
                                "image_url": f"data:image/jpeg;base64,{base64_img}",
                            },
                            include_image_base64=False,
                        )

                        return " ".join(
                            page.markdown
                            for page in ocr_response.pages
                            if page.markdown
                        )

                    with concurrent.futures.ThreadPoolExecutor(
                        max_workers=10
                    ) as executor:
                        results = list(executor.map(ocr_image, images))

                    return clean_text(" ".join(results))

            else:
                raise ValueError(f"OCR not supported for type: {doc_type}")

            combined_text = " ".join(
                page.markdown for page in pages if getattr(page, "markdown", None)
            )
            return clean_text(combined_text)

        except Exception as e:
            raise ValueError(f"Failed to OCR {doc_type.upper()}: {e}")

    def extract_csv_text(content: bytes) -> str:
        text_rows = []
        try:
            decoded = content.decode("utf-8", errors="ignore")
            reader = csv.reader(io.StringIO(decoded))
            rows = list(reader)

            if not rows:
                return ""

            headers = [header.strip() for header in rows[0]]

            for idx, row in enumerate(rows[1:], start=2):
                if not any(cell.strip() for cell in row):
                    continue

                cells = []
                for i, cell in enumerate(row):
                    key = headers[i] if i < len(headers) else f"Column{i+1}"
                    value = cell.strip()
                    if value:
                        cells.append(f"{key}: {value}")

                if cells:
                    context_line = f"[CSV Row {idx}] " + " | ".join(cells)
                    text_rows.append(context_line)

        except Exception as e:
            raise ValueError(f"Failed to parse CSV: {e}")

        return clean_text(" ".join(text_rows))

    def extract_html_text(content: bytes) -> str:
        try:
            html = content.decode("utf-8", errors="ignore")
            soup = BeautifulSoup(html, "lxml")

            for tag in soup(["script", "style", "noscript", "template"]):
                tag.decompose()

            main = soup.find(["main", "article"]) or soup.body
            text = (
                main.get_text(separator=" ", strip=True)
                if main
                else soup.get_text(separator=" ", strip=True)
            )
            return clean_text(text)
        except Exception as e:
            raise ValueError(f"Failed to parse HTML: {e}")

    try:
        if doc_type == "pdf":
            return extract_pdf_text(content)
        elif doc_type == "docx":
            return extract_docx_text(content)
        elif doc_type == "email":
            return extract_email_text(content)
        elif doc_type == "excel":
            return extract_excel_text(content)
        elif doc_type in {"jpg", "png", "ppt", "pptx"}:
            return extract_with_mistral_ocr(content, doc_type, url)
        elif doc_type == "csv":
            return extract_csv_text(content)
        elif doc_type == "html":
            return extract_html_text(content)
        else:
            return f"No context available for {doc_type.upper()}"
    except Exception as e:
        raise ValueError(f"Failed to parse {doc_type.upper()}: {e}")


def chunk_sections(text: str, max_tokens: int, overlap_tokens: int) -> List[dict]:
    """
    Split text into overlapping chunks using LangChain splitter
    Uses tokenizer for precise token counting to optimize embedding performance
    """
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=max_tokens,
        chunk_overlap=overlap_tokens,
        length_function=lambda text: len(tokenizer.encode(text, truncation=False)),
        separators=["\n\n", "\n", ". ", "! ", "? ", " ", ", ", ""],
    )
    chunks = splitter.split_text(text)
    return [
        {"text": chunk, "tokens": len(tokenizer.encode(chunk, truncation=False))}
        for chunk in chunks
    ]


async def process_document(url: str, force_refresh: bool) -> dict:
    """
    Main document processing pipeline
    1. Detect document type
    2. Download and extract content
    3. Chunk text into optimized segments
    4. Cache results in Redis for future use
    """
    # Step 1: Identify document type from URL
    doc_type = identify_document_type(url)
    if not doc_type:
        # Return default response for unsupported formats
        return {
            "status": "processed",
            "doc_hash": "invalid_item",
            "doc_type": "unknown",
            "average_tokens": 0,
            "max_tokens": 10,
            "min_tokens": 0,
            "no_of_chunks": 5,
            "chunks": [
                "No Context Available",
                "No Context Available",
                "No Context Available",
                "No Context Available",
                "No Context Available",
            ],
        }

    # Step 2: Download document content
    content = download_document(url)
    if not content:
        return {
            "status": "processed",
            "doc_hash": "invalid_item",
            "doc_type": doc_type,
            "average_tokens": 0,
            "max_tokens": 10,
            "min_tokens": 0,
            "no_of_chunks": 5,
            "chunks": [
                "No Context Available",
                "No Context Available",
                "No Context Available",
                "No Context Available",
                "No Context Available",
            ],
        }

    # Generate unique hash for document content
    doc_hash = hashlib.md5(content).hexdigest()
    redis_key = f"doc:{doc_hash}"

    # Check if document is already cached (unless force refresh)
    if client.exists(redis_key) and not force_refresh:
        cached_data = client.hgetall(redis_key)
        if cached_data:
            return {
                "status": "cached",
                "doc_hash": doc_hash,
                "doc_type": doc_type,
                "average_tokens": float(cached_data["average_tokens"]),
                "max_tokens": int(cached_data["max_tokens"]),
                "min_tokens": int(cached_data["min_tokens"]),
                "no_of_chunks": int(cached_data.get("no_of_chunks", 0)),
            }

    # Step 3: Extract text content using format-specific parser
    text = parse_document_content(content, doc_type, url)

    if not text:
        return {
            "status": "processed",
            "doc_hash": doc_hash,
            "doc_type": doc_type,
            "average_tokens": 0,
            "max_tokens": 10,
            "min_tokens": 0,
            "no_of_chunks": 5,
            "chunks": [
                "No Context Available",
                "No Context Available",
                "No Context Available",
                "No Context Available",
                "No Context Available",
            ],
        }

    # Step 4: Chunk text into optimal segments for embedding
    max_tokens = int(os.getenv("DOCUMENT_CHUNK_SIZE"))
    overlap_tokens = int(os.getenv("DOCUMENT_CHUNK_OVERLAP"))
    chunk_infos = chunk_sections(text, max_tokens, overlap_tokens)
    token_counts = [info["tokens"] for info in chunk_infos]
    chunks = [info["text"] for info in chunk_infos]

    # Step 5: Cache document metadata in Redis
    client.hset(
        redis_key,
        mapping={
            "doc_hash": doc_hash,
            "doc_type": doc_type,
            "average_tokens": sum(token_counts) / len(token_counts),
            "max_tokens": max(token_counts),
            "min_tokens": min(token_counts),
            "no_of_chunks": len(chunks),
        },
    )

    return {
        "status": "processed",
        "doc_hash": doc_hash,
        "doc_type": doc_type,
        "average_tokens": sum(token_counts) / len(token_counts),
        "max_tokens": max(token_counts),
        "min_tokens": min(token_counts),
        "no_of_chunks": len(chunks),
        "chunks": chunks,
    }
